import os
import openpyxl
from openpyxl.styles import PatternFill

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set a title for the worksheet
ws.title = "Sample Sheet"

# Write data to cells
ws['A1'] = "Hello"
ws['B1'] = "World!"

# Define the file path
file_path = r'C:\Users\plmala\OneDrive - Coloplast A S\Desktop\Python\Internal Trainers\SampleExcelWorkbook.xlsx'

# Ensure the directory exists
os.makedirs(os.path.dirname(file_path), exist_ok=True)

# Apply colors to specific cells
ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Fill color for B1
ws['B1'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Save the workbook
wb.save(file_path)
print(f"Workbook saved at {file_path}")
